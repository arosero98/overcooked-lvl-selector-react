#!/usr/bin/env python3
"""
Diagnostic script to inspect the structure of the Excel file.
"""

import openpyxl
from pathlib import Path


def inspect_excel(filepath, max_rows=85, max_cols=30):
    """Inspect and display the structure of the Excel file"""
    wb = openpyxl.load_workbook(filepath)
    worlds=[1]
    for i in worlds:
        sheet = wb[f"World {i}"]

        print(f"Sheet name: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print(f"Max column: {sheet.max_column}")
        print("\n" + "=" * 80)
        print("First few rows and columns:")
        print("=" * 80 + "\n")

        # Display the first max_rows rows and max_cols columns
        for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, max_col=max_cols), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                value = cell.value if cell.value is not None else ""
                # Truncate long values
                value_str = str(value)[:30]
                row_data.append(f"[{col_idx}] {value_str}")

            print(f"Row {row_idx:3d}: {' | '.join(row_data)}")

            if row_idx % 10 == 0:
                print("-" * 80)


def main():
    excel_file = Path("Overcooked 2 Full Task Analysis.xlsx")

    if not excel_file.exists():
        print(f"Error: File '{excel_file}' not found!")
        return

    print(f"Inspecting {excel_file}...\n")
    inspect_excel(excel_file)


if __name__ == "__main__":
    main()
