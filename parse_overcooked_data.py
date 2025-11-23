#!/usr/bin/env python3
"""
Script to parse Overcooked 2 Full Task Analysis.xlsx and generate JSON data.
"""

import openpyxl
import json
import re
from pathlib import Path
from datetime import time


# Mapping of Excel field names to JSON keys
FIELD_MAPPING = {
    # Basic level info
    "Level Time (mm:ss)": "time_to_complete",
    "Does Clock start at \"Go\" ?": "start_at_go",
    "Is Environment Fixed?": "fixed_environment",
    "Number of Recipes": "composite_num",  # We'll handle variations separately
    "Dish Washer?": "dish_washer",
    "Score for 1 Star (2 Players)": "one_star_score",
    "Number of Dishes": "num_dishes",
    "Challenge Score": "challenge_score",

    # Points and scoring
    "Points given for completing Composite Challenge 1": "points_composite_1",
    "Points given for completing Composite Challenge 2": "points_composite_2",
    "Tip Multiplier (x1)": "tip_multiplier_x1",
    "Tip Multiplier (x2)": "tip_multiplier_x2",
    "Tip Multiplier (x3)": "tip_multiplier_x3",

    # Recipe and challenge info
    "Is Recipe (Composite Challenge) Order Fixed?": "recipe_order_fixed",
    "Number of Completed Composite Challenges needed to achieve 1 Star (2 Players)": "num_completed_composite_for_star",
    "Penalty for failing a Composite Challenge": "penalty_failed_composite",

    # Equipment
    "Chopping Board?": "has_chopping_board",
    "Number of Chopping Boards": "num_chopping_boards",
    "Oven?": "has_oven",
    "Number of Ovens": "num_ovens",
    "Stove-Tops?": "has_stove_tops",
    "Number of Stove-Tops?": "num_stove_tops",
    "Mixers?": "has_mixers",
    "Number of Mixers?": "num_mixers",

    # Notes
    "Additional Notes": "additional_notes",
}


def is_level_header(value):
    """Check if a value is a level header like 'Level 1.1', 'Level 2.3', etc."""
    if value is None:
        return False
    value_str = str(value).strip()
    # Match pattern like "Level 1.1", "Level 2.3", etc.
    match = re.match(r'^Level (\d+\.\d+)$', value_str, re.IGNORECASE)
    return match.group(1) if match else None


def format_level_key(level_num):
    """Convert level number like '1.1' to 'Level_1_1'"""
    return f"Level_{level_num.replace('.', '_')}"


def convert_time_to_mmss(value):
    """Convert Excel time format to mm:ss"""
    if value is None:
        return "manual-input-needed"

    # If it's already a datetime.time object from Excel
    if isinstance(value, time):
        # Excel stores time as HH:MM:SS where HH=minutes, MM=seconds for game time
        minutes = value.hour
        seconds = value.minute
        return f"{minutes}:{seconds:02d}"

    # If it's a string
    value_str = str(value).strip()

    # Try to parse HH:MM:SS format (Excel format where HH=minutes, MM=seconds)
    time_match = re.match(r'^(\d+):(\d+):(\d+)$', value_str)
    if time_match:
        hours, mins, secs = map(int, time_match.groups())
        # Hours field represents minutes, minutes field represents seconds
        minutes = hours
        seconds = mins
        return f"{minutes}:{seconds:02d}"

    # Try mm:ss format (already correct)
    if re.match(r'^\d+:\d{2}$', value_str):
        return value_str

    return "manual-input-needed"


def parse_number_of_recipes(value):
    """Parse 'Number of Recipes' which might be '1 (2 variations)' or just '1.0'"""
    if value is None or str(value).strip() == "":
        return None, None

    value_str = str(value).strip()

    # Check for pattern like "1 (2 variations)"
    match = re.match(r'^(\d+)\s*\((\d+)\s*variations?\)$', value_str, re.IGNORECASE)
    if match:
        composite = match.group(1)
        variations = match.group(2)
        return composite, variations

    # Otherwise just a number
    try:
        num = float(value_str)
        if num.is_integer():
            return str(int(num)), None
        return str(num), None
    except (ValueError, AttributeError):
        return None, None


def validate_and_clean_value(field_name, value):
    """Validate and clean values based on field type. Returns None if value is invalid/missing."""
    if value is None or str(value).strip() == "":
        return None

    value_str = str(value).strip()

    # Time format validation
    if field_name == "time_to_complete":
        result = convert_time_to_mmss(value)
        return None if result == "manual-input-needed" else result

    # Yes/No fields
    yes_no_fields = [
        "start_at_go", "fixed_environment", "dish_washer",
        "has_chopping_board", "has_oven", "has_stove_tops", "has_mixers",
        "recipe_order_fixed"
    ]
    if field_name in yes_no_fields:
        lower_val = value_str.lower()
        if lower_val in ["yes", "no", "y", "n"]:
            return "yes" if lower_val in ["yes", "y"] else "no"
        return None

    # Numeric fields
    numeric_fields = [
        "composite_num", "variation_num", "one_star_score", "challenge_score", "num_dishes",
        "points_composite_1", "points_composite_2",
        "tip_multiplier_x1", "tip_multiplier_x2", "tip_multiplier_x3",
        "num_completed_composite_for_star",
        "num_chopping_boards", "num_ovens", "num_stove_tops", "num_mixers"
    ]
    if field_name in numeric_fields:
        # Skip "NA" values
        if value_str.upper() == "NA":
            return None

        try:
            num_val = float(value_str)
            if num_val.is_integer():
                return str(int(num_val))
            return str(num_val)
        except (ValueError, AttributeError):
            return None

    # Text fields (penalty, notes) - accept as-is if not empty
    if field_name in ["penalty_failed_composite", "additional_notes"]:
        return value_str if value_str else None

    return value_str


def parse_sheet(sheet, sheet_name):
    """Parse a single sheet and extract level data"""
    levels_data = {}

    # Find all level headers in row 1
    print(f"\n{'='*60}")
    print(f"Processing {sheet_name}...")
    print('='*60)
    print("Scanning for level headers in row 1...")
    level_columns = {}

    for cell in sheet[1]:
        level_num = is_level_header(cell.value)
        if level_num:
            level_key = format_level_key(level_num)
            level_columns[level_key] = {
                'header_col': cell.column,
                'field_col': cell.column + 1,
                'value_col': cell.column + 2
            }
            print(f"Found {cell.value} at column {cell.column} (field col: {cell.column + 1}, value col: {cell.column + 2})")

    print(f"\nFound {len(level_columns)} levels")

    # For each level, scan through rows to find field values
    for level_key, cols in level_columns.items():
        print(f"\nProcessing {level_key}...")

        # Initialize level data with default values (ordered as they appear in Excel)
        levels_data[level_key] = {
            "time_to_complete": "",                     # Row 15
            "start_at_go": "",                          # Row 16
            "one_star_score": "",                       # Row 17
            "points_composite_1": "",                   # Row 18
            "points_composite_2": "",                   # Row 19
            "tip_multiplier_x1": "",                    # Row 20
            "tip_multiplier_x2": "",                    # Row 21
            "tip_multiplier_x3": "",                    # Row 22
            "fixed_environment": "",                    # Row 25
            "recipe_order_fixed": "",                   # Row 26
            "composite_num": "",                        # Row 27
            "variation_num": "",                        # Row 27 (variations)
            "dish_washer": "",                          # Row 28
            "num_dishes": "",                           # Row 29
            "has_chopping_board": "",                   # Row 30
            "num_chopping_boards": "",                  # Row 31
            "has_oven": "",                             # Row 32
            "num_ovens": "",                            # Row 33
            "has_stove_tops": "",                       # Row 34
            "num_stove_tops": "",                       # Row 35
            "has_mixers": "",                           # Row 36
            "num_mixers": "",                           # Row 37
            "num_completed_composite_for_star": "",     # Row 60
            "penalty_failed_composite": "",             # Row 61
            "challenge_score": "",                      # Row 64
            "additional_notes": "",                     # Row 66
            "video_link": ""                            # Not in Excel
        }

        field_col = cols['field_col']
        value_col = cols['value_col']

        # Scan through all rows in the field column
        for row in range(1, sheet.max_row + 1):
            field_cell = sheet.cell(row=row, column=field_col)
            field_value = field_cell.value

            if field_value:
                field_str = str(field_value).strip()

                # Check if this is one of our target fields
                if field_str in FIELD_MAPPING:
                    json_key = FIELD_MAPPING[field_str]
                    value_cell = sheet.cell(row=row, column=value_col)
                    value = value_cell.value

                    # Special handling for "Number of Recipes"
                    if field_str == "Number of Recipes":
                        composite, variations = parse_number_of_recipes(value)
                        if composite:
                            levels_data[level_key]["composite_num"] = composite
                            print(f"  composite_num: {composite}")
                        if variations:
                            levels_data[level_key]["variation_num"] = variations
                            print(f"  variation_num: {variations}")
                    else:
                        cleaned_value = validate_and_clean_value(json_key, value)
                        # Set value (empty string if None)
                        levels_data[level_key][json_key] = cleaned_value if cleaned_value is not None else ""
                        if cleaned_value is not None:
                            print(f"  {json_key}: {cleaned_value}")

        # Post-processing: If has_X is "no", set num_X to "0"
        equipment_mappings = [
            ("has_chopping_board", "num_chopping_boards"),
            ("has_oven", "num_ovens"),
            ("has_stove_tops", "num_stove_tops"),
            ("has_mixers", "num_mixers"),
        ]

        for has_field, num_field in equipment_mappings:
            if levels_data[level_key].get(has_field) == "no":
                if levels_data[level_key].get(num_field) == "":
                    levels_data[level_key][num_field] = "0"

    return levels_data


def parse_excel_file(filepath):
    """Parse the Excel file and extract level data from all World sheets"""
    wb = openpyxl.load_workbook(filepath)
    all_levels_data = {}

    # Parse World 1 through World 6
    for world_num in range(1, 7):
        sheet_name = f"World {world_num}"

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"\nWarning: Sheet '{sheet_name}' not found, skipping...")
            continue

        sheet = wb[sheet_name]
        levels_data = parse_sheet(sheet, sheet_name)

        # Merge into all_levels_data
        all_levels_data.update(levels_data)

    return all_levels_data


def main():
    """Main function"""
    excel_file = Path("Overcooked 2 Full Task Analysis.xlsx")
    output_file = Path("overcooked_levels_data.json")

    if not excel_file.exists():
        print(f"Error: File '{excel_file}' not found!")
        return

    # Load existing JSON to preserve video_link values
    existing_data = {}
    if output_file.exists():
        print(f"Loading existing {output_file} to preserve video_link values...\n")
        with open(output_file, 'r', encoding='utf-8') as f:
            existing_data = json.load(f)

    print(f"Parsing {excel_file}...\n")

    try:
        levels_data = parse_excel_file(excel_file)

        # Preserve video_link from existing data
        for level_key in levels_data:
            if level_key in existing_data and "video_link" in existing_data[level_key]:
                levels_data[level_key]["video_link"] = existing_data[level_key]["video_link"]

        # Write to JSON file
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(levels_data, f, indent=2, ensure_ascii=False)

        print(f"\n✓ Successfully parsed {len(levels_data)} levels")
        print(f"✓ Data written to {output_file}")

        # Print summary
        print("\nSummary:")
        for level_key in sorted(levels_data.keys(), key=lambda x: [int(n) for n in x.replace('Level_', '').split('_')]):
            field_count = len(levels_data[level_key]) - 1  # Subtract 1 for video_link
            print(f"  {level_key}: {field_count} fields parsed")

    except Exception as e:
        print(f"Error parsing file: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
